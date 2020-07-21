import argparse
import json
import os
from datetime import datetime
from dateutil import parser as dateparser, relativedelta
from boxsdk import JWTAuth, Client, LoggingClient
from openpyxl import Workbook
from pprint import pprint

is_parent_folder = True
current_enterprise_id = None

# Events types LOGIN and at least one File event is required
event_types='LOGIN,UPLOAD,DOWNLOAD,PREVIEW,DELETE,COPY,EDIT,MOVE,SHARE'

# Limit of Box events to retrieve before starting to paginate
limit = 250

# Dictionaries to store data
folder_collaborations_dict = {}
events_dict = {}
last_login_dict = {}

# Main function
def main(box_config, parent_folder_id, day_lookback, start_date_str, end_date_str):
    # Get the Box service account client
    auth = JWTAuth.from_settings_file(box_config)
    client = Client(auth)
    current_enterprise = client.get_current_enterprise()
    global current_enterprise_id
    current_enterprise_id = current_enterprise.id

    service_account = client.user().get()
    print('Found Service Account with name: {0}, id: {1}, and login: {2}'.format(service_account.name, service_account.id, service_account.login))

    start_date = None
    end_date = None
    if day_lookback is not None:
        # Get the current date and the date for one month ago
        start_date = end_date - relativedelta.relativedelta(days=day_lookback)
        end_date = datetime.utcnow()
    elif (start_date_str is not None) and (end_date_str is not None):
        date_format = '%Y-%m-%d'
        start_date = datetime.strptime(start_date_str, date_format)
        end_date = datetime.strptime(end_date_str, date_format)
    else:
        raise Exception('--day_lookback OR --start_date AND --end_date are required parameters!')
    print('Using date range for events - start date: {0} and end date: {1}'.format(start_date, end_date))

    # Create a collaboration dictionary
    traverse_folder_tree(client, parent_folder_id)
    print('Found collab count: {0}'.format(len(folder_collaborations_dict)))

    # Get Box events
    get_box_events(client, limit, 0, start_date, end_date)
    print('Found event count: {0}'.format(len(events_dict)))

    # Generate Excel report
    create_excel_report()
    print('Finished!')

# Function to traverse a folder hierachy, get associated collaborations,
def traverse_folder_tree(client, parent_folder_id):
    global is_parent_folder
    if is_parent_folder:
        is_parent_folder = False
        parent_folder = client.folder(folder_id=parent_folder_id).get(fields=['id', 'type', 'name', 'path_collection'])
        get_folder_collaborations(client, parent_folder)
    else:
        # Get parent folder and its direct descendants
        # TODO: Implement marker-based pagination
        items = client.folder(folder_id=parent_folder_id).get_items(fields=['id', 'type', 'name', 'path_collection'])

        # Loop through the folder children
        for item in items:
            get_folder_collaborations(client, item)


# Fucntion to get item collaborations
def get_folder_collaborations(client, item):
    # Check if the item type is a folder or a file
    collaborations = None
    if item.type == 'folder':
        # Get the collaborations on the folder
        # TODO: Implement marker-based pagination
        collaborations = client.folder(folder_id=item.id).get_collaborations(fields=['id', 'name', 'role', 'created_at', 'created_by', 'accessible_by', 'status', 'acknowledged_at', 'invite_email'])

        # Parse the collaboration values so we can then update the folder collab dictionary
        parse_collaboration_values(client, collaborations, item)

        # We found a folder and therefore should call the current function relfectively
        traverse_folder_tree(client, item.id)



# Function to create a dictionary to store collaborations
def parse_collaboration_values(client, collaborations, item):
    # Create a string for the item path
    path = ''
    id_path = ''
    for path_segment in item.path_collection['entries']:
        path += '/{0}'.format(path_segment.name)
        id_path += '/{0}'.format(path_segment.id)
    path += '/{0}'.format(item.name)
    id_path = '/{0}'.format(item.id)
    print('Found item with id: {0}, type: {1}, and path: {2}'.format(item.id, item.type, path))

    # Loop through the collabations
    for collab in collaborations:
        collab_created_by = None
        if collab.created_by is not None:
            collab_created_by = collab.created_by.login

        # Check if the collaboration accessible type is a group or a user since Groups will not have a login
        accessible_by = collab.accessible_by
        if accessible_by is not None:
            if accessible_by.type != 'group':
                user_type = None

                # Check if the user is a Service Account
                if accessible_by.login.endswith('@boxdevedition.com') and accessible_by.login.startswith('AutomationUser_'):
                    user_type = 'Service Account'
                # Check if the user is an App User
                elif accessible_by.login.endswith('@boxdevedition.com') and accessible_by.login.startswith('AppUser_'):
                    user_type = 'App User'
                # Else we have a managed user or an external user
                else:
                    # Check if the collaboration status = pending. You cannot get the user objecct for an external uesr if the collaboration has not been accepted.
                    if collab.status != 'pending':
                        # Get the user parameters
                        user = client.user(user_id=accessible_by.id).get(fields=['id', 'name', 'login', 'enterprise'])

                        # Check if the user enterprise is null
                        if user.enterprise is not None:
                            # If the user EID is equal to the current EID, then its a managed user.
                            if user.enterprise.id == current_enterprise_id:
                                user_type = 'Managed'
                            # Else, we have an external user
                            else:
                                user_type = 'External'
                        else:
                            user_type = ''

                # Call function to add an value to the folder collaboration dictionary
                update_folder_collab_dict(collab, item, path, id_path, user_type, accessible_by.id, accessible_by.name, accessible_by.login, collab_created_by)
            else:
                # We found a group so we need to get the group members
                group_memberships = client.group(group_id=accessible_by.id).get_memberships(fields=['user', 'group'])

                # For each group member, add a item to the folder collaboration dictionary
                for membership in group_memberships:
                    # Call function to add an value to the folder collaboration dictionary
                    update_folder_collab_dict(collab, item, path, id_path, 'Group: {0}'.format(membership.group.name), membership.user.id, membership.user.name, membership.user.login, collab_created_by)
        else:
            # Call function to add an value to the folder collaboration dictionary
            update_folder_collab_dict(collab, item, path, id_path, 'Pending User', None, None, collab.invite_email, collab_created_by)

# Update the folder collaboration dictionary
def update_folder_collab_dict(collab, item, path, id_path, collab_type, accessible_by_id, accessible_by_name, accessible_by_login, collab_created_by):
    print('Found collaboration with id: {0}, for item: {1}, collaborator id: {2}, and collaborator login: {3}'.format(collab.id, item.name, accessible_by_id, accessible_by_login))
    # Popuplate the folder collabation dictionary
    collab_key = '{0}.{1}.{2}'.format(collab.id, item.id, accessible_by_id)
    folder_collaborations_dict[collab_key] = {}
    folder_collaborations_dict[collab_key]['item_path'] = path
    folder_collaborations_dict[collab_key]['item_id_path'] = id_path
    folder_collaborations_dict[collab_key]['item_name'] = item.name
    folder_collaborations_dict[collab_key]['item_id'] = item.id
    folder_collaborations_dict[collab_key]['item_type'] = item.type
    folder_collaborations_dict[collab_key]['collab_name'] = accessible_by_name
    folder_collaborations_dict[collab_key]['collab_login'] = accessible_by_login
    folder_collaborations_dict[collab_key]['collab_type'] = collab_type
    folder_collaborations_dict[collab_key]['collab_role'] = collab.role
    folder_collaborations_dict[collab_key]['collab_status'] = collab.status
    folder_collaborations_dict[collab_key]['collab_created_by_login'] = collab_created_by
    folder_collaborations_dict[collab_key]['collab_invite_date'] = collab.created_at
    folder_collaborations_dict[collab_key]['collab_acknowledged_date'] = collab.acknowledged_at


# Function to get Box events
def get_box_events(client, limit, stream_position, created_after, created_before):
    chunk_size = None
    while chunk_size != 0:
        # Populate the URL query parameters
        url_params = 'stream_type=admin_logs&event_type={0}&limit={1}&stream_position={2}&created_after={3}&created_before={4}'.format(event_types, limit, stream_position, created_after, created_before)

        # GET request to retrieve events
        events_response = client.make_request(
            'GET',
            client.get_url('events?{0}'.format(url_params)),
        ).json()

        # Get the next stream position
        stream_position = events_response['next_stream_position']
        chunk_size = events_response['chunk_size']
        print('Found events response with chunk_size={0}, next_stream_position={1}'.format(chunk_size, stream_position))

        # Loop through the events and store them in a dictionary.
        events = events_response['entries']
        for event in events:
            event_id = event['event_id']
            events_dict[event_id] = event

# Function to create an excel workbook
def create_excel_report():
    print('Creating Excel report...')

    # Create an excel workbook
    workbook = Workbook()
    current_time = datetime.now()
    report_filename = 'collab_report_{0}.xlsx'.format(current_time.strftime("%Y_%m"))
    worksheet = workbook.active

    # Create header row
    row_count = 1
    worksheet.cell(column=1, row=row_count, value='Path')
    worksheet.cell(column=2, row=row_count, value='Path IDs')
    worksheet.cell(column=3, row=row_count, value='Item Name')
    worksheet.cell(column=4, row=row_count, value='Item ID')
    worksheet.cell(column=5, row=row_count, value='Item Type')
    worksheet.cell(column=6, row=row_count, value='Collaborator Name')
    worksheet.cell(column=7, row=row_count, value='Collaborator Login')
    worksheet.cell(column=8, row=row_count, value='Collaborator Type')
    worksheet.cell(column=9, row=row_count, value='Collaborator Permission')
    worksheet.cell(column=10, row=row_count, value='Status')
    worksheet.cell(column=11, row=row_count, value='Inviter Email')
    worksheet.cell(column=12, row=row_count, value='Invited Date')
    worksheet.cell(column=13, row=row_count, value='Invite Accepted Date')
    worksheet.cell(column=14, row=row_count, value='Last Login Date')
    worksheet.cell(column=15, row=row_count, value='Last Activity Type')
    worksheet.cell(column=16, row=row_count, value='Last Activity User')
    worksheet.cell(column=17, row=row_count, value='IP Address')
    worksheet.cell(column=18, row=row_count, value='Last Activity Date')
    row_count += 1

    # Loop through the folder collaboration dictionary and populate the excel cells
    print('Iterating through a total of {0} collaborations:'.format(len(folder_collaborations_dict)))
    for key, value in folder_collaborations_dict.items():
        item_path = value['item_path']
        item_id_path = value['item_id_path']
        item_id = value['item_id']
        item_name = value['item_name']
        item_type = value['item_type']
        collab_name = value['collab_name']
        collab_login = value['collab_login']
        collab_type = value['collab_type']
        collab_role = value['collab_role']
        collab_status = value['collab_status']
        collab_created_by_login = value['collab_created_by_login']
        collab_invite_date = value['collab_invite_date']
        collab_acknowledged_date = value['collab_acknowledged_date']

        worksheet.cell(column=1, row=row_count, value=item_path)
        worksheet.cell(column=2, row=row_count, value=item_id_path)
        worksheet.cell(column=3, row=row_count, value=item_name)
        worksheet.cell(column=4, row=row_count, value=item_id)
        worksheet.cell(column=5, row=row_count, value=item_type)
        worksheet.cell(column=6, row=row_count, value=collab_name)
        worksheet.cell(column=7, row=row_count, value=collab_login)
        worksheet.cell(column=8, row=row_count, value=collab_type)
        worksheet.cell(column=9, row=row_count, value=collab_role)
        worksheet.cell(column=10, row=row_count, value=collab_status)
        worksheet.cell(column=11, row=row_count, value=collab_created_by_login)
        worksheet.cell(column=12, row=row_count, value=collab_invite_date)
        worksheet.cell(column=13, row=row_count, value=collab_acknowledged_date)

        # Call the get_last_login function to get the last login
        last_login = get_last_login(collab_login)
        print('{0} - Adding last login datetime: {1}'.format(row_count-1, last_login))
        worksheet.cell(column=14, row=row_count, value=last_login)

        # Get the last file event details
        last_file_event = get_last_file_event(key, collab_login, item_type, item_id)
        if last_file_event:
            print('{0} - Adding last event: {1}'.format(row_count-1, last_file_event['event_type']))
            worksheet.cell(column=15, row=row_count, value=last_file_event['event_type'])
            worksheet.cell(column=16, row=row_count, value=last_file_event['created_by_login'])
            worksheet.cell(column=17, row=row_count, value=last_file_event['ip_address'])
            worksheet.cell(column=18, row=row_count, value=last_file_event['created_at'])
        row_count += 1

    # Save the workbook to the local file system
    # TODO: Implement the ability to optionally upload the file to a folder in Box
    workbook.save(filename = report_filename)

# Function to get the last login date given a login email
def get_last_login(login):
    # Check if the last login date already exists in the dictionary
    if login in last_login_dict:
        return last_login_dict[login]
    else:
        # Filter the events dictionary where the login is in found in source.login
        filtered_events = { key:value for (key,value) in events_dict.items() if value['event_type'] == 'LOGIN' and login in value['created_by']['login'] }
        if filtered_events:
            # Since events are sequential, pop the last login event item
            key, value = filtered_events.popitem()

            # Store the last login in the last login dictionary
            created_at = value['created_at']
            created_by_login = value['created_by']['login']
            last_login_dict[created_by_login] = created_at
            return last_login_dict[created_by_login]
        else:
            # Service Accounts, groups, or the users that have not acepted a collaboration invite will not have a last login date
            return ''

# Function to get the last file activity in a parent folder
def get_last_file_event(collab_id, login, item_type, item_id):

    filtered_events = {}
    if item_type == 'folder':
        # Filter the events dictionary where the item_id is in found in source.parent.id and for a specific login
        for key, value in events_dict.items():
            if(value['event_type'] != 'LOGIN' and login == value['created_by']['login'] and item_id in value['source']['parent']['id']):
                filtered_events[key] = value
    else:
        # Filter the events dictionary where the item_id is in found in source.id and for a specific login
        for key, value in events_dict.items():
            if(value['event_type'] != 'LOGIN' and login == value['created_by']['login'] and item_id in value['source']['item_id']):
                filtered_events[key] = value

    if filtered_events:
        # Since events are sequential, pop the last login event item
        key, value = filtered_events.popitem()

        # Store the last file event in the last_file_activity_dict dictionary
        created_at = value['created_at']
        created_by_login = value['created_by']['login']
        event_type = value['event_type']
        ip_address = value['ip_address']
        return { 'event_type': event_type, 'created_by_login': created_by_login, 'created_at': created_at, 'ip_address': ip_address }
    else:
        # Service Accounts, groups, or the users that have not acepted a collaboration invite will not have a last login date
        return ''

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Create a Last Login Report')
    parser.add_argument('--box_config', metavar='/path/to/my/box_config.json', required=True, help='The path to your Box JWT app config')
    parser.add_argument('--parent_folder_id', metavar='12345679', required=True, help='Parent Folder ID to begin searching for collaborations')
    parser.add_argument('--day_lookback', metavar='1', type=int, help='Integer that represents the amount of days to look back for events')
    parser.add_argument('--start_date', metavar='2020-01-10', help='String representing the start date to look for events. Format is YYYY-MM-DD')
    parser.add_argument('--end_date', metavar='2020-07-15', help='String representing the end date to look for events. Format is YYYY-MM-DD')

    args = parser.parse_args()

    main(args.box_config, args.parent_folder_id, args.day_lookback, args.start_date, args.end_date)
